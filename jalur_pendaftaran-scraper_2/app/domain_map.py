import os
import openpyxl
import tldextract
from .utils import load_json, save_json

def build_domain_map(path: str, id_col: str, url_col: str, state_dir: str) -> dict:
    cache_path = os.path.join(state_dir, "domain_map.json")
    if os.path.exists(cache_path):
        return load_json(cache_path, default={})

    wb = openpyxl.load_workbook(path)
    ws = wb.active

    header = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if v:
            header[str(v).strip()] = c

    if id_col not in header or url_col not in header:
        raise RuntimeError(f"Kolom {id_col=} atau {url_col=} tidak ada di {path}")

    out = {}
    for r in range(2, ws.max_row + 1):
        uid = ws.cell(r, header[id_col]).value
        url = ws.cell(r, header[url_col]).value
        if not uid or not url:
            continue
        ext = tldextract.extract(str(url))
        dom = f"{ext.domain}.{ext.suffix}"
        if dom:
            try:
                out[dom] = int(uid)
            except Exception:
                continue

    save_json(cache_path, out)
    return out
