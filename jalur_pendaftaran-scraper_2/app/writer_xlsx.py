import openpyxl
import os
import time

class PricesWorkbookWriter:
    def __init__(self, template_path: str, out_path: str):
        self.template_path = template_path
        self.out_path = out_path

        # kalau output sudah ada, lanjutkan (resume output)
        if os.path.exists(out_path):
            self.wb = openpyxl.load_workbook(out_path)
        else:
            self.wb = openpyxl.load_workbook(template_path)

        self.ws = self.wb["Format Excel"]
        self.headers = [self.ws.cell(1, c).value for c in range(1, self.ws.max_column + 1)]

    def prepare_fresh_sheet(self, fresh: bool = True):
        # kalau output sudah ada, jangan hapus isi (biar resume)
        if not fresh:
            return
        if self.ws.max_row > 1:
            self.ws.delete_rows(2, self.ws.max_row - 1)

    def append_items(self, items: list[dict]):
        for it in items:
            # fallback supaya tidak kosong total
            if it.get("priceable_id") is None:
                it["priceable_id"] = 0

            row = []
            for h in self.headers:
                row.append(it.get(h))
            self.ws.append(row)

    def save(self):
        self.wb.save(self.out_path)

    def autosave(self, min_interval_s: int = 10):
        # autosave sederhana agar tidak terlalu sering I/O
        now = time.time()
        last = getattr(self, "_last_save", 0)
        if now - last >= min_interval_s:
            self.save()
            self._last_save = now